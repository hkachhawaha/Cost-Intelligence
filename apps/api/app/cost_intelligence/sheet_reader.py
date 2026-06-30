"""Google Sheets reader (ingestion layer). Reads a PUBLIC workbook via its xlsx export — no
OAuth; the sheet must be shared "anyone with the link can view".

Each Nexus tab carries a human title banner in row 1 and the real column headers in row 2, so
`parse_workbook` skips row 1, treats row 2 as the header, and stops at the first row whose key
column is blank. `parse_workbook` (bytes → dict) is pure and unit-tested against a fixture;
`read` adds the network fetch.
"""

from __future__ import annotations

import io
import re
from typing import Any

import httpx

from app.core.config import settings

# The "Read Me" cover tab carries scenario prose, not tabular data — never ingested.
SKIP_TABS = {"Read Me"}
_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")


class SheetReadError(RuntimeError):
    """The workbook could not be fetched or parsed (bad URL, not public, empty)."""


def extract_spreadsheet_id(url_or_id: str) -> str:
    """Accept a full Google Sheets URL or a bare spreadsheet id."""
    if not url_or_id:
        raise SheetReadError("no spreadsheet URL provided")
    m = _ID_RE.search(url_or_id)
    if m:
        return m.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", url_or_id):
        return url_or_id
    raise SheetReadError(f"could not extract a spreadsheet id from {url_or_id!r}")


def _xlsx_export_url(spreadsheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def parse_workbook(data: bytes) -> dict[str, list[dict[str, Any]]]:
    """Parse an xlsx workbook → {tab_name: [row_dict, ...]} (header is row 2, data from row 3,
    'Read Me' skipped, rows truncated at the first blank key column)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface a clean domain error
        raise SheetReadError(f"workbook is not a valid xlsx: {exc}") from exc

    out: dict[str, list[dict[str, Any]]] = {}
    for name in wb.sheetnames:
        if name in SKIP_TABS:
            continue
        ws = wb[name]
        header: list[str] | None = None
        rows: list[dict[str, Any]] = []
        for idx, raw in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue  # row 1 = title banner
            if idx == 1:
                header = [
                    str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(raw)
                ]
                continue
            if header is None:
                break
            # Stop at the first row whose key (first) column is empty — end of data.
            if not raw or raw[0] in (None, ""):
                break
            row = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
            rows.append(row)
        if header:
            out[name] = rows
    wb.close()
    if not out:
        raise SheetReadError("workbook contained no ingestable tabs")
    return out


class GoogleSheetReader:
    def __init__(self, timeout: int | None = None):
        self.timeout = timeout or settings.ci_sheet_fetch_timeout_s

    async def fetch_xlsx(self, spreadsheet_id: str) -> bytes:
        url = _xlsx_export_url(spreadsheet_id)
        async with httpx.AsyncClient(timeout=self.timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            if resp.status_code != 200:
                raise SheetReadError(
                    f"could not fetch workbook (HTTP {resp.status_code}); is the sheet shared "
                    f"'anyone with the link can view'?"
                )
            return resp.content

    async def read(self, url_or_id: str) -> dict[str, list[dict[str, Any]]]:
        sid = extract_spreadsheet_id(url_or_id)
        data = await self.fetch_xlsx(sid)
        return parse_workbook(data)
