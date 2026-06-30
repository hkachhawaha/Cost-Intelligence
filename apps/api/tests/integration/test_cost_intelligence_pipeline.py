"""Cost Intelligence integration (realignment) — live Postgres, migration 011.

Connect → Agent Memory snapshot stored + status; Refresh increments the version and serves from
memory. The Google Sheet read is monkeypatched to a tiny in-memory workbook (no network), so the
whole ingest → relationships → insights → memory pipeline is exercised deterministically.

Run: RUN_DB_TESTS=1 uv run pytest apps/api/tests/integration/test_cost_intelligence_pipeline.py -v
"""

from __future__ import annotations

import asyncio
import io

import pytest

from tests.conftest import requires_db

pytestmark = requires_db

_CONTRACT_HDR = [
    "Contract_ID", "Vendor_Name", "Category", "Subcategory", "Region", "Contract_Value_USD",
    "Annual_Value_USD", "Effective_Date", "Expiration_Date", "Renewal_Notice_Days", "Auto_Renew",
    "Pricing_Model", "Payment_Terms", "Rebate_Clause", "SLA_Penalty_Clause", "Volume_Commitment",
    "Most_Recent_Amendment", "Internal_Owner", "Department", "Contract_Status",
]
_SPEND_HDR = [
    "Transaction_ID", "Transaction_Date", "Vendor_Name", "Contract_ID", "PO_Number",
    "Cost_Center", "Department", "GL_Account", "Description", "Amount_USD", "Payment_Method",
    "Invoice_Reference", "Fiscal_Quarter",
]


def _fixture_xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    contracts = wb.create_sheet("Contracts")
    contracts.append(["NEXUS — Contract Register (banner)"])
    contracts.append(_CONTRACT_HDR)
    contracts.append(["NXC-1", "Acme Cloud Services", "Cloud", "CDN", "Global", 400000, 200000,
                      "2024-01-01", "2026-01-01", 30, "Y", "Consumption", "Net-30", "N", "N",
                      "None", "None", "J. Doe", "IT", "Active"])
    spend = wb.create_sheet("Spend Ledger")
    spend.append(["NEXUS — AP Spend Ledger (banner)"])
    spend.append(_SPEND_HDR)
    spend.append(["TXN-1", "2025-06-01", "Acme Cloud Services", "NXC-1", "PO-1", "CC-1", "IT",
                  "GL-1", "svc", 260000, "ACH", "INV-1", "Q2 2025"])  # matched, over ACV
    spend.append(["TXN-2", "2025-06-02", "Rogue Vendor Ltd", "", "", "CC-9", "Ops", "GL-9",
                  "misc", 50000, "ACH", "INV-9", "Q2 2025"])  # maverick
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _dsn() -> str:
    from app.core.config import settings

    return settings.database_url_sync.replace("postgresql+psycopg://", "postgresql://")


@pytest.fixture()
def clean_ci():
    import psycopg

    def _truncate():
        with psycopg.connect(_dsn(), autocommit=True) as conn, conn.cursor() as cur:
            cur.execute("TRUNCATE ci_memory_snapshot, ci_data_source CASCADE")

    _truncate()  # start from a clean version counter (global max)
    yield
    _truncate()


def _patch_reader(monkeypatch, data: bytes) -> None:
    from app.cost_intelligence.sheet_reader import GoogleSheetReader

    async def _fake_fetch(self, spreadsheet_id):  # noqa: ANN001
        return data

    monkeypatch.setattr(GoogleSheetReader, "fetch_xlsx", _fake_fetch)


_URL = "https://docs.google.com/spreadsheets/d/FAKEID0000000000000000000/edit"


def test_connect_stores_memory_and_status(clean_ci, monkeypatch):
    from app.core.database import SessionFactory
    from app.cost_intelligence.service import CostIntelligenceService

    _patch_reader(monkeypatch, _fixture_xlsx())

    async def run():
        async with SessionFactory() as s:
            svc = CostIntelligenceService(s)
            res = await svc.connect(_URL, "Nexus Test")
            await s.commit()
            status = await svc.status()
            snap = await svc.snapshot()
            return res, status, snap

    res, status, snap = asyncio.run(run())
    assert res["status"] == "connected" and res["memory_version"] == 1
    assert status["connected"] is True and status["total_records"] == 3  # 1 contract + 2 spend
    assert snap is not None and len(snap["contracts"]) == 1 and len(snap["spend"]) == 2
    # Pipeline ran: maverick + overspend surfaced; KPIs present.
    types = {o["type"] for o in snap["opportunities"]}
    assert "Maverick spend" in types and "Overspend vs ACV" in types
    assert snap["kpis"]["total"] == 310000.0
    assert snap["relationships"]["counts"]["maverickRecords"] == 1


def test_refresh_increments_version(clean_ci, monkeypatch):
    from app.core.database import SessionFactory
    from app.cost_intelligence.service import CostIntelligenceService

    _patch_reader(monkeypatch, _fixture_xlsx())

    async def run():
        async with SessionFactory() as s:
            svc = CostIntelligenceService(s)
            await svc.connect(_URL, "Nexus Test")
            await s.commit()
            r2 = await svc.refresh()
            await s.commit()
            return r2

    r2 = asyncio.run(run())
    assert r2["memory_version"] == 2  # refresh wrote a new snapshot version


def test_test_connection_does_not_store(clean_ci, monkeypatch):
    from app.core.database import SessionFactory
    from app.cost_intelligence.service import CostIntelligenceService

    _patch_reader(monkeypatch, _fixture_xlsx())

    async def run():
        async with SessionFactory() as s:
            svc = CostIntelligenceService(s)
            probe = await svc.test_connection(_URL)
            await s.commit()
            status = await svc.status()
            return probe, status

    probe, status = asyncio.run(run())
    assert probe["ok"] is True
    assert probe["tabs"] == {"Contracts": 1, "Spend Ledger": 2}
    assert status["status"] == "never"  # test connection never persists
