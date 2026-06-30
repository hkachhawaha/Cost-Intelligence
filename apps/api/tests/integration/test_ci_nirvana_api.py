"""NirvanAI API integration (realignment Phase 6) — live Postgres, migration 011.

POST /ci/nirvana/ask answers from the connected Agent Memory (deterministic when no LLM key);
asking before any data is connected returns 404. Sheet read is monkeypatched (no network).

Run: RUN_DB_TESTS=1 uv run pytest apps/api/tests/integration/test_ci_nirvana_api.py -v
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from tests.conftest import requires_db

pytestmark = requires_db

_URL = "https://docs.google.com/spreadsheets/d/NIRVANAFIXTURE0000000000/edit"


def _xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    c = wb.create_sheet("Contracts")
    c.append(["banner"])
    c.append(["Contract_ID", "Vendor_Name", "Annual_Value_USD", "Expiration_Date",
              "Renewal_Notice_Days", "Auto_Renew", "Contract_Status"])
    c.append(["NXC-1", "Acme Cloud", 200000, "2025-07-15", 30, "Y", "Active"])
    s = wb.create_sheet("Spend Ledger")
    s.append(["banner"])
    s.append(["Transaction_ID", "Transaction_Date", "Vendor_Name", "Contract_ID", "PO_Number",
              "Amount_USD", "Invoice_Reference"])
    s.append(["TXN-1", "2025-03-01", "Acme Cloud", "NXC-1", "PO-1", 260000, "INV-1"])
    s.append(["TXN-2", "2025-03-02", "Rogue Vendor", "", "", 50000, "INV-9"])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _dsn() -> str:
    from app.core.config import settings

    return settings.database_url_sync.replace("postgresql+psycopg://", "postgresql://")


def _truncate():
    import psycopg
    from redis import Redis

    from app.core.config import settings

    with psycopg.connect(_dsn(), autocommit=True) as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE ci_memory_snapshot, ci_data_source CASCADE")
    # Also clear the Agent Memory warm cache so an empty DB really reads as empty.
    r = Redis.from_url(str(settings.redis_url))
    r.delete("ci:memory:latest")
    r.close()


@pytest.fixture()
def client(monkeypatch):
    from app.core import auth as auth_mod
    from app.core.database import SessionFactory, get_session
    from app.cost_intelligence.sheet_reader import GoogleSheetReader
    from app.main import app

    async def _fake_fetch(self, sid):  # noqa: ANN001
        return _xlsx()

    async def _noop():
        return None

    monkeypatch.setattr(GoogleSheetReader, "fetch_xlsx", _fake_fetch)
    monkeypatch.setattr(auth_mod.jwks_cache, "_refresh", _noop)

    async def _session():
        s = SessionFactory()
        try:
            yield s
            await s.commit()
        finally:
            await s.close()

    app.dependency_overrides[get_session] = _session
    _truncate()
    yield TestClient(app)
    app.dependency_overrides.clear()
    _truncate()


def test_nirvana_ask_before_connect_returns_404(client):
    r = client.post("/api/v1/ci/nirvana/ask", json={"question": "what is recoverable?"})
    assert r.status_code == 404


def test_nirvana_ask_after_connect_answers_from_memory(client):
    assert client.post("/api/v1/ci/data-source/connect",
                       json={"url": _URL, "name": "Nirvana"}).status_code == 200
    r = client.post("/api/v1/ci/nirvana/ask", json={"question": "What is recoverable right now?"})
    assert r.status_code == 200
    body = r.json()
    assert body["answer"] and body["source"] in {"deterministic", "llm"}
    # Grounded: the overspend on NXC-1 (260k vs 200k ACV) is recoverable → a $ figure is quoted.
    assert "$" in body["answer"]


def test_nirvana_ask_benchmark_is_refused(client):
    client.post("/api/v1/ci/data-source/connect", json={"url": _URL, "name": "Nirvana"})
    r = client.post("/api/v1/ci/nirvana/ask", json={"question": "Are we paying above market?"})
    assert r.status_code == 200
    assert "external market data" in r.json()["answer"]
